from __future__ import annotations

import asyncio
import base64
import io
import json
import mimetypes
import re
import shutil
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from uuid import uuid4

import fitz
from docx import Document
from openpyxl import load_workbook
from PIL import Image

from backend.app.config import settings


IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
TEXT_SUFFIXES = {".txt", ".md", ".json", ".pdf", ".docx", ".xlsx"}
ALLOWED_ATTACHMENT_SUFFIXES = IMAGE_SUFFIXES | TEXT_SUFFIXES


@dataclass
class ResolvedAttachments:
    text: str
    images: list[str]
    items: list[dict[str, Any]]


class AttachmentStore:
    """Persistent local chat attachments with bounded text extraction."""

    def __init__(self) -> None:
        self.root = settings.root_dir / "data" / "uploads" / "chat"
        self.root.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def validate_session_id(session_id: str) -> str:
        if not re.fullmatch(r"[A-Za-z0-9_-]{1,96}", session_id):
            raise ValueError("会话标识不合法")
        return session_id

    @staticmethod
    def validate_attachment_id(attachment_id: str) -> str:
        if not re.fullmatch(r"[a-f0-9]{32}", attachment_id):
            raise ValueError("附件标识不合法")
        return attachment_id

    def _session_dir(self, session_id: str) -> Path:
        return self.root / self.validate_session_id(session_id)

    @staticmethod
    def _public_meta(meta: dict[str, Any]) -> dict[str, Any]:
        session_id = str(meta["session_id"])
        attachment_id = str(meta["id"])
        return {
            "id": attachment_id,
            "name": str(meta["name"]),
            "content_type": str(meta["content_type"]),
            "size": int(meta["size"]),
            "kind": str(meta["kind"]),
            "url": f"/api/attachments/{attachment_id}?session_id={session_id}",
        }

    async def save(
        self,
        *,
        session_id: str,
        filename: str,
        content_type: str | None,
        data: bytes,
    ) -> dict[str, Any]:
        return await asyncio.to_thread(
            self._save_sync,
            session_id=session_id,
            filename=filename,
            content_type=content_type,
            data=data,
        )

    def _save_sync(
        self,
        *,
        session_id: str,
        filename: str,
        content_type: str | None,
        data: bytes,
    ) -> dict[str, Any]:
        session_dir = self._session_dir(session_id)
        session_dir.mkdir(parents=True, exist_ok=True)
        safe_name = Path(filename).name or "attachment.bin"
        suffix = Path(safe_name).suffix.lower()
        if suffix not in ALLOWED_ATTACHMENT_SUFFIXES:
            raise ValueError(f"不支持的聊天附件类型：{suffix or '未知'}")
        attachment_id = uuid4().hex
        file_path = session_dir / f"{attachment_id}{suffix}"
        file_path.write_bytes(data)
        is_image = suffix in IMAGE_SUFFIXES
        extracted_text = "" if is_image else self._extract_text(file_path, suffix)
        meta = {
            "id": attachment_id,
            "session_id": session_id,
            "name": safe_name,
            "suffix": suffix,
            "content_type": content_type or mimetypes.guess_type(safe_name)[0] or "application/octet-stream",
            "size": len(data),
            "kind": "image" if is_image else "document",
            "extracted_text": extracted_text[:24000],
        }
        (session_dir / f"{attachment_id}.json").write_text(
            json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return self._public_meta(meta)

    def _extract_text(self, path: Path, suffix: str) -> str:
        if suffix in {".txt", ".md"}:
            return path.read_text(encoding="utf-8", errors="ignore")[:24000]
        if suffix == ".json":
            try:
                value = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
                return json.dumps(value, ensure_ascii=False, indent=2)[:24000]
            except json.JSONDecodeError:
                return path.read_text(encoding="utf-8", errors="ignore")[:24000]
        if suffix == ".pdf":
            pdf = fitz.open(path)
            try:
                pieces = [pdf[index].get_text("text") for index in range(min(pdf.page_count, 30))]
                return "\n\n".join(pieces)[:24000]
            finally:
                pdf.close()
        if suffix == ".docx":
            document = Document(path)
            return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())[:24000]
        if suffix == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                rows: list[str] = []
                for row in sheet.iter_rows(min_row=1, max_row=120, values_only=True):
                    rows.append("\t".join(str(value or "") for value in row[:12]))
                return "\n".join(rows)[:24000]
            finally:
                workbook.close()
        return ""

    def _load_meta(self, session_id: str, attachment_id: str) -> tuple[dict[str, Any], Path]:
        session_dir = self._session_dir(session_id)
        attachment_id = self.validate_attachment_id(attachment_id)
        meta_path = session_dir / f"{attachment_id}.json"
        if not meta_path.exists():
            raise FileNotFoundError(f"附件 {attachment_id} 不存在或不属于当前会话")
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        file_path = session_dir / f"{attachment_id}{meta['suffix']}"
        if not file_path.exists():
            raise FileNotFoundError(f"附件文件 {attachment_id} 已丢失")
        return meta, file_path

    async def resolve(self, session_id: str, attachment_ids: list[str]) -> ResolvedAttachments:
        return await asyncio.to_thread(self._resolve_sync, session_id, attachment_ids)

    def _resolve_sync(self, session_id: str, attachment_ids: list[str]) -> ResolvedAttachments:
        images: list[str] = []
        document_image_count = 0
        text_parts: list[str] = []
        items: list[dict[str, Any]] = []
        for attachment_id in attachment_ids[: settings.max_chat_attachments]:
            meta, file_path = self._load_meta(session_id, attachment_id)
            items.append(self._public_meta(meta))
            if meta["kind"] == "image":
                images.append(self._image_base64(file_path))
            else:
                if meta.get("extracted_text"):
                    text_parts.append(
                        f"[附件：{meta['name']}]\n{meta['extracted_text']}"
                    )
                remaining = max(
                    0, settings.max_chat_document_images - document_image_count
                )
                if remaining:
                    document_images = self._document_images(
                        file_path, str(meta["suffix"]), remaining
                    )
                    images.extend(document_images)
                    document_image_count += len(document_images)
        return ResolvedAttachments(
            text="\n\n".join(text_parts)[:32000],
            images=images,
            items=items,
        )

    @staticmethod
    def _image_base64(path: Path) -> str:
        with Image.open(path) as image:
            image.load()
            if max(image.size) <= 1800 and path.stat().st_size <= 4 * 1024 * 1024:
                return base64.b64encode(path.read_bytes()).decode("ascii")
            image.thumbnail((1800, 1800))
            buffer = io.BytesIO()
            if image.mode not in {"RGB", "L"}:
                image = image.convert("RGB")
            image.save(buffer, format="JPEG", quality=90, optimize=True)
            return base64.b64encode(buffer.getvalue()).decode("ascii")

    @staticmethod
    def _document_images(path: Path, suffix: str, limit: int) -> list[str]:
        """Render document visuals so formulas and circuit diagrams reach the VLM."""

        if suffix == ".pdf":
            document = fitz.open(path)
            try:
                rendered: list[str] = []
                for page_index in range(min(document.page_count, limit)):
                    pixmap = document[page_index].get_pixmap(
                        matrix=fitz.Matrix(1.35, 1.35), alpha=False
                    )
                    rendered.append(
                        base64.b64encode(pixmap.tobytes("png")).decode("ascii")
                    )
                return rendered
            finally:
                document.close()
        if suffix not in {".docx", ".xlsx"}:
            return []

        prefix = "word/media/" if suffix == ".docx" else "xl/media/"
        rendered: list[str] = []
        try:
            with zipfile.ZipFile(path) as archive:
                members = sorted(
                    member
                    for member in archive.infolist()
                    if member.filename.startswith(prefix)
                    and not member.is_dir()
                    and member.file_size <= 8 * 1024 * 1024
                )
                for member in members[:limit]:
                    try:
                        with Image.open(io.BytesIO(archive.read(member))) as image:
                            image.load()
                            image.thumbnail((1800, 1800))
                            if image.mode not in {"RGB", "L"}:
                                image = image.convert("RGB")
                            buffer = io.BytesIO()
                            image.save(buffer, format="JPEG", quality=88, optimize=True)
                            rendered.append(
                                base64.b64encode(buffer.getvalue()).decode("ascii")
                            )
                    except (OSError, ValueError):
                        continue
        except (OSError, zipfile.BadZipFile):
            return []
        return rendered

    def file_for_response(self, session_id: str, attachment_id: str) -> tuple[dict[str, Any], Path]:
        return self._load_meta(session_id, attachment_id)

    def list_public(self, session_id: str) -> list[dict[str, Any]]:
        session_dir = self._session_dir(session_id)
        if not session_dir.exists():
            return []
        items: list[tuple[float, dict[str, Any]]] = []
        for meta_path in session_dir.glob("*.json"):
            try:
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
                file_path = session_dir / f"{meta['id']}{meta['suffix']}"
                if file_path.is_file():
                    items.append((meta_path.stat().st_mtime, self._public_meta(meta)))
            except (KeyError, OSError, TypeError, ValueError, json.JSONDecodeError):
                continue
        return [item for _, item in sorted(items, key=lambda value: value[0])]

    def enrich_history(
        self, session_id: str, messages: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        """Restore attachment URLs, including legacy messages that only kept names."""

        available = self.list_public(session_id)
        by_id = {str(item["id"]): item for item in available}
        by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for item in available:
            by_name[str(item["name"])].append(item)
        used_legacy_ids: set[str] = set()
        restored: list[dict[str, Any]] = []
        for original in messages:
            message = dict(original)
            stored_attachments = message.get("attachments")
            resolved: list[dict[str, Any]] = []
            if isinstance(stored_attachments, list):
                message.pop("attachments", None)
                for stored in stored_attachments:
                    if not isinstance(stored, dict):
                        continue
                    current = by_id.get(str(stored.get("id", "")))
                    if current:
                        resolved.append(current)
            if not resolved and message.get("role") == "user":
                content = str(message.get("content", ""))
                marker = re.search(r"(?:\n)?\[附件：([^\]]+)\]\s*$", content)
                if marker:
                    for name in (part.strip() for part in marker.group(1).split("、")):
                        candidate = next(
                            (
                                item for item in by_name.get(name, [])
                                if str(item["id"]) not in used_legacy_ids
                            ),
                            None,
                        )
                        if candidate:
                            resolved.append(candidate)
                            used_legacy_ids.add(str(candidate["id"]))
                    if resolved:
                        message["content"] = content[:marker.start()].rstrip()
            if resolved:
                message["attachments"] = resolved
            restored.append(message)
        return restored

    async def delete_session(self, session_id: str) -> bool:
        return await asyncio.to_thread(self._delete_session_sync, session_id)

    def _delete_session_sync(self, session_id: str) -> bool:
        root = self.root.resolve()
        target = self._session_dir(session_id).resolve()
        if target.parent != root:
            raise ValueError("附件会话目录不合法")
        if not target.exists():
            return False
        shutil.rmtree(target)
        return True
